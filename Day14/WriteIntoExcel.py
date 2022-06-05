import openpyxl
import os

# Writing single data
# file = os.getcwd()+"//write.xlsx"
# workbook = openpyxl.load_workbook(file)
# sheet = workbook.active
#
# for r in range(1,5):
#     for c in range(1,4):
#         sheet.cell(r,c).value = "Welcome"
#
# workbook.save(file)

# Multiple Data
file = os.getcwd()+"//write.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

sheet.cell(5,1).value = "Amaresh"
sheet.cell(5,2).value = "Kavadimatti"

sheet.cell(6,1).value = "Raichur"
sheet.cell(6,2).value = "Mudgal"

sheet.cell(7,1).value = "India"
sheet.cell(7,2).value = "Karnataka"

workbook.save(file)

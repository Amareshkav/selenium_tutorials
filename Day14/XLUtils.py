import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowno, colno).value


def writeData(file, sheetName, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowno, colno).value = data
    workbook.save(file)


def fillGreenColor(file, sheetName, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    green_fill = PatternFill(start_color='60b212',
                             end_color='60b212',
                             fill_type='solid'
                             )
    sheet.cell(rowno, colno).fill = green_fill
    workbook.save(file)


def fillRedColor(file, sheetName, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    red_fill = PatternFill(start_color='ff0000',
                           end_color='ff0000',
                           fill_type='solid'
                           )
    sheet.cell(rowno, colno).fill = red_fill
    workbook.save(file)
